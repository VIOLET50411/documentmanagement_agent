"""Excel/CSV parser with sheet-aware structured output."""

from __future__ import annotations

import csv
from pathlib import Path

MOJIBAKE_MARKERS = (
    "\u951f",
    "\u9359",
    "\u7487",
    "\u9200\u20ac",
    "\ufffd",
    "\u00c3",
    "\u00e6",
    "\u00e7",
    "\u00b5\u00c4",
    "\u00c9\u00e8\u00b1\u00b8",
)


class ExcelParser:
    """Parse CSV or spreadsheet files into structured markdown-like elements."""

    def parse(self, file_path: str) -> list[dict]:
        path = Path(file_path)
        if path.suffix.lower() == ".csv":
            return self._parse_csv(path)
        return self._parse_xlsx(path)

    def _parse_xlsx(self, path: Path) -> list[dict]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return self._parse_xlsx_with_pandas(path)

        workbook = load_workbook(filename=path, data_only=False, read_only=True)
        elements: list[dict] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = self._extract_rows_from_sheet(worksheet)
            if not rows:
                continue
            elements.append(
                {
                    "type": "table",
                    "text": self._render_markdown_table(rows),
                    "metadata": {
                        "sheet": sheet_name,
                        "section_title": sheet_name,
                        "row_count": max(len(rows) - 1, 0),
                        "column_count": max((len(row) for row in rows), default=0),
                        "merged_ranges": [],
                        "parser": "openpyxl",
                    },
                }
            )
            elements.append(self._build_sheet_overview_element(sheet_name=sheet_name, rows=rows, parser="openpyxl"))
        return elements

    def _parse_xlsx_with_pandas(self, path: Path) -> list[dict]:
        try:
            import pandas as pd
        except ImportError:
            return [
                {
                    "type": "table",
                    "text": f"文件 {path.name} 需要 pandas/openpyxl 才能完成表格解析。",
                    "metadata": {"sheet": path.stem, "parser": "fallback"},
                }
            ]

        sheets = pd.read_excel(path, sheet_name=None)
        elements: list[dict] = []
        for sheet_name, frame in sheets.items():
            clean_frame = frame.fillna("")
            rows = [list(map(str, clean_frame.columns.tolist()))]
            rows.extend([list(map(str, row)) for row in clean_frame.values.tolist()[:500]])
            if len(rows) == 1 and not any(rows[0]):
                continue
            elements.append(
                {
                    "type": "table",
                    "text": self._render_markdown_table(rows),
                    "metadata": {
                        "sheet": sheet_name,
                        "section_title": sheet_name,
                        "row_count": len(rows) - 1,
                        "column_count": max((len(row) for row in rows), default=0),
                        "parser": "pandas",
                    },
                }
            )
            elements.append(self._build_sheet_overview_element(sheet_name=sheet_name, rows=rows, parser="pandas"))
        return elements

    def _parse_csv(self, path: Path) -> list[dict]:
        text = self._decode_csv_text(path)
        reader = csv.reader(text.splitlines())
        rows = list(reader)
        if not rows:
            return []
        visible_rows = rows[:501]
        return [
            {
                "type": "table",
                "text": self._render_markdown_table(visible_rows),
                "metadata": {
                    "sheet": path.stem,
                    "section_title": path.stem,
                    "row_count": max(len(rows) - 1, 0),
                    "column_count": max((len(row) for row in visible_rows), default=0),
                    "parser": "csv",
                },
            },
            self._build_sheet_overview_element(sheet_name=path.stem, rows=visible_rows, parser="csv"),
        ]

    def _decode_csv_text(self, path: Path) -> str:
        raw = path.read_bytes()
        for encoding in ("utf-8-sig", "utf-8"):
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError:
                continue

        best_text = ""
        best_score = float("-inf")
        for encoding in ("gb18030", "gbk"):
            try:
                decoded = raw.decode(encoding)
            except UnicodeDecodeError:
                continue
            score = self._score_decoded_text(decoded)
            if score > best_score:
                best_score = score
                best_text = decoded
        if best_text:
            return best_text
        return raw.decode("utf-8", errors="replace")

    def _score_decoded_text(self, text: str) -> float:
        penalty = sum(text.count(marker) for marker in MOJIBAKE_MARKERS)
        cjk = sum(1 for char in text if "\u4e00" <= char <= "\u9fff")
        printable = sum(1 for char in text if char.isprintable())
        return cjk * 4 + printable - penalty * 10

    def _extract_rows_from_sheet(self, worksheet) -> list[list[str]]:
        rows: list[list[str]] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
            rendered_row = [self._cell_to_text(cell) for cell in row]
            if any(value for value in rendered_row):
                rows.append(rendered_row)
            if row_index >= 500:
                break
        return rows

    def _render_markdown_table(self, rows: list[list[str]]) -> str:
        normalized_rows = rows or [[]]
        width = max((len(row) for row in normalized_rows), default=0)
        header = (normalized_rows[0] + [""] * width)[:width]
        body = [((row or []) + [""] * width)[:width] for row in normalized_rows[1:]]
        rendered = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * width) + " |",
        ]
        for row in body:
            rendered.append("| " + " | ".join(row) + " |")
        return "\n".join(rendered)

    def _cell_to_text(self, cell) -> str:
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value)

    def _build_sheet_overview_element(self, *, sheet_name: str, rows: list[list[str]], parser: str) -> dict:
        header = [cell for cell in (rows[0] if rows else []) if cell][:6]
        row_count = max(len(rows) - 1, 0)
        column_count = max((len(row) for row in rows), default=0)
        header_text = "、".join(header) if header else "未识别列名"
        return {
            "type": "paragraph",
            "text": f"工作表《{sheet_name}》概览：共{row_count}行、{column_count}列，字段包括{header_text}。",
            "metadata": {
                "sheet": sheet_name,
                "section_title": sheet_name,
                "row_count": row_count,
                "column_count": column_count,
                "parser": f"{parser}_overview",
            },
        }
